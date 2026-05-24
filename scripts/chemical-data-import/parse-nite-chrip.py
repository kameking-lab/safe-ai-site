#!/usr/bin/env python3
"""
NITE-CHRIP 統合版 GHS 分類結果 (list_nite_all.xlsx) のパーサ

  $ python3 scripts/chemical-data-import/parse-nite-chrip.py

入力 : scripts/chemical-data-import/tmp/list_nite_all.xlsx
出力 : web/src/data/chemicals-nite/classifications.jsonl
       web/src/data/chemicals-nite/manifest.json

xlsx は政府(独立行政法人 NITE) の公開データ。出典: 元 URL
https://www.chem-info.nite.go.jp/chem/ghs/files/list_nite_all.xlsx
本セッションは直接ダウンロード不可だったため、GitHub ミラー
https://github.com/Ameyanagi/risk_assessment_list/raw/main/reference/list_nite_all.xlsx
経由で取得 (cached_at: 2026-03-25)。

GHS 区分値は短縮コード化 (rep-byte 削減目的):
  "区分1A"  -> "1A"
  "区分1B"  -> "1B"
  "区分1"   -> "1"
  "区分2"   -> "2"
  "区分3"   -> "3"
  "区分4"   -> "4"
  "区分5"   -> "5"
  "液化ガス"    -> "L"
  "圧縮ガス"    -> "P"
  "深冷液化ガス" -> "D"
  "溶解ガス"    -> "S"
  "区分に該当しない（分類対象外）" -> "N"
  "区分外"   -> "X"
  "分類できない" -> "U"
その他は素通し。Node 側で逆引きする辞書を持つ。
"""

import hashlib
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl  # type: ignore
except ImportError:
    print("[parse-nite-chrip] openpyxl 未インストール: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "scripts" / "chemical-data-import" / "tmp" / "list_nite_all.xlsx"
OUT_DIR = ROOT / "web" / "src" / "data" / "chemicals-nite"
OUT_JSONL = OUT_DIR / "classifications.jsonl"
OUT_MANIFEST = OUT_DIR / "manifest.json"

SOURCE_URL = "https://www.chem-info.nite.go.jp/chem/ghs/files/list_nite_all.xlsx"
MIRROR_URL = (
    "https://github.com/Ameyanagi/risk_assessment_list/raw/main/reference/list_nite_all.xlsx"
)
SHEET_NAME = "NITE統合版GHS分類結果"

# 元 xlsx の列見出し → 出力 JSON キー (snake -> camel ではなく短いキーで JSONL を軽量化)
COLUMN_TO_KEY = {
    "爆発物": "explosives",
    "可燃性ガス": "flammableGases",
    "エアゾール": "aerosols",
    "酸化性ガス": "oxidizingGases",
    "高圧ガス": "gasesUnderPressure",
    "引火性液体": "flammableLiquids",
    "可燃性固体": "flammableSolids",
    "自己反応性化学品": "selfReactive",
    "自然発火性液体": "pyrophoricLiquids",
    "自然発火性固体": "pyrophoricSolids",
    "自己発熱性化学品": "selfHeating",
    "水反応可燃性化学品": "waterReactive",
    "酸化性液体": "oxidizingLiquids",
    "酸化性固体": "oxidizingSolids",
    "有機過酸化物": "organicPeroxides",
    "金属腐食性化学品": "corrosiveToMetals",
    "鈍性化爆発物": "desensitizedExplosives",
    "急性毒性（経口）": "acuteToxOral",
    "急性毒性（経皮）": "acuteToxDermal",
    "急性毒性（吸入：ガス）": "acuteToxInhalGas",
    "急性毒性（吸入：蒸気）": "acuteToxInhalVapor",
    "急性毒性（吸入：粉塵、ミスト）": "acuteToxInhalDust",
    "皮膚腐食性／刺激性": "skinCorrIrr",
    "眼に対する重篤な損傷性／眼刺激性": "eyeDamageIrr",
    "呼吸器感作性": "respSens",
    "皮膚感作性": "skinSens",
    "生殖細胞変異原性": "mutagen",
    "発がん性": "carcinogen",
    "生殖毒性": "reproTox",
    "特定標的臓器毒性（単回暴露）": "stotSingle",
    "特定標的臓器毒性（反復暴露）": "stotRepeat",
    "誤えん有害性": "aspiration",
    "水生環境有害性　短期（急性）": "aquaticAcute",
    "水生環境有害性　長期（慢性）": "aquaticChronic",
    "オゾン層への有害性": "ozone",
}

# GHS 区分値 → 短縮コード
CATEGORY_SHORTHAND = {
    "区分1A": "1A",
    "区分1B": "1B",
    "区分1C": "1C",
    "区分1": "1",
    "区分2": "2",
    "区分2A": "2A",
    "区分2B": "2B",
    "区分3": "3",
    "区分4": "4",
    "区分5": "5",
    "液化ガス": "L",
    "圧縮ガス": "P",
    "深冷液化ガス": "D",
    "溶解ガス": "S",
    "区分に該当しない（分類対象外）": "N",
    "区分に該当しない": "N",
    "区分外": "X",
    "分類できない": "U",
    "": None,
}


def shorten(value):
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    return CATEGORY_SHORTHAND.get(text, text)


def is_classified(short):
    """『区分N』に該当 (実害指摘あり) かどうか"""
    if not short:
        return False
    if short in ("N", "X", "U", "L", "P", "D", "S"):
        return False
    return True


def sha256_of(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def main():
    if not SRC.exists():
        print(f"[parse-nite-chrip] 入力ファイル不在: {SRC}", file=sys.stderr)
        print(
            f"  curl -sL -o {SRC} {MIRROR_URL}\n  (元URL: {SOURCE_URL})",
            file=sys.stderr,
        )
        sys.exit(1)

    sha = sha256_of(SRC)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(
            f"[parse-nite-chrip] シート不在: {SHEET_NAME} / available={wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(2)
    ws = wb[SHEET_NAME]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    if not header or header[0] != "CAS RN":
        print(f"[parse-nite-chrip] ヘッダ不正: {header[:3]}", file=sys.stderr)
        sys.exit(3)

    # ヘッダ → 列インデックス
    col_idx = {h: i for i, h in enumerate(header) if h}
    required = ["CAS RN", "物質名称", "GHS分類結果_ID"]
    for r in required:
        if r not in col_idx:
            print(f"[parse-nite-chrip] 必須列不在: {r}", file=sys.stderr)
            sys.exit(4)

    detail_url_col = col_idx.get("詳細情報（NITE HP　GHS分類結果）")
    label_url_col = col_idx.get("モデルラベル掲載ページURL")
    sds_url_col = col_idx.get("モデルＳＤＳ掲載ページURL")

    cas_col = col_idx["CAS RN"]
    name_col = col_idx["物質名称"]
    id_col = col_idx["GHS分類結果_ID"]

    written = 0
    skipped = 0
    classified_only = 0
    with open(OUT_JSONL, "w", encoding="utf-8") as out:
        for row in rows_iter:
            if not row or row[cas_col] in (None, ""):
                skipped += 1
                continue
            cas = str(row[cas_col]).strip()
            name = (str(row[name_col]).strip() if row[name_col] else "")
            ghs_id = (str(row[id_col]).strip() if row[id_col] else "")

            ghs = {}
            n_classified = 0
            for jp_col, key in COLUMN_TO_KEY.items():
                if jp_col not in col_idx:
                    continue
                raw = row[col_idx[jp_col]]
                short = shorten(raw)
                if short is None:
                    continue
                ghs[key] = short
                if is_classified(short):
                    n_classified += 1

            if n_classified > 0:
                classified_only += 1

            entry = {
                "cas": cas,
                "nameJa": name,
                "ghsId": ghs_id,
                "ghs": ghs,
                "classifiedCount": n_classified,
            }
            if detail_url_col is not None and row[detail_url_col]:
                entry["chripUrl"] = str(row[detail_url_col]).strip()
            if label_url_col is not None and row[label_url_col]:
                entry["modelLabelUrl"] = str(row[label_url_col]).strip()
            if sds_url_col is not None and row[sds_url_col]:
                entry["modelSdsUrl"] = str(row[sds_url_col]).strip()

            out.write(json.dumps(entry, ensure_ascii=False) + "\n")
            written += 1

    manifest = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "source": {
            "officialUrl": SOURCE_URL,
            "mirrorUrl": MIRROR_URL,
            "sha256": sha,
            "fileSize": SRC.stat().st_size,
            "sheetName": SHEET_NAME,
        },
        "license": "政府データ(独立行政法人 製品評価技術基盤機構 NITE)、出典明記のうえ自由利用可",
        "totalSubstances": written,
        "skippedRows": skipped,
        "withAnyClassifiedHazard": classified_only,
        "shorthandCodes": {
            "1A/1B/1C/1/2/2A/2B/3/4/5": "GHS 区分N (実害指摘あり)",
            "L": "液化ガス",
            "P": "圧縮ガス",
            "D": "深冷液化ガス",
            "S": "溶解ガス",
            "N": "区分に該当しない（分類対象外）",
            "X": "区分外",
            "U": "分類できない",
        },
    }
    with open(OUT_MANIFEST, "w", encoding="utf-8") as mf:
        json.dump(manifest, mf, ensure_ascii=False, indent=2)
        mf.write("\n")

    print(f"[parse-nite-chrip] wrote {written} substances → {OUT_JSONL.relative_to(ROOT)}")
    print(f"[parse-nite-chrip] manifest → {OUT_MANIFEST.relative_to(ROOT)}")
    print(f"[parse-nite-chrip] skipped={skipped} withClassifiedHazard={classified_only}")


if __name__ == "__main__":
    main()
